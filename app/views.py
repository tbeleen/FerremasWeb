import random
import string
from django.shortcuts import render, redirect
from django.urls import reverse
import requests
from django.contrib import messages
from django.views.decorators.csrf import csrf_exempt
from django.views.decorators.csrf import csrf_protect
from django.views.decorators.http import require_POST
from django.http import JsonResponse
from django.core.files.storage import FileSystemStorage
import json
from reportlab.pdfgen import canvas
from reportlab.lib.pagesizes import letter
from reportlab.lib.units import mm
from reportlab.lib import colors
from django.http import HttpResponse
import os
from ferremas import settings
from django.core.paginator import Paginator
import openpyxl
from datetime import date, timedelta, datetime


def index(request):

    marcas_response = requests.get('http://localhost:3000/marca')
    marca = marcas_response.json()
    response = requests.get('http://localhost:3000/productos')  
    productos = response.json()

    paginator = Paginator(productos, 8)  
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    response = requests.get('http://localhost:3000/productos')
    productos = response.json()

    return render(request, 'app/index.html', {
        'productos': productos,
        'productos': page_obj,
        'page_obj': page_obj,
        'marcas': marca
})

from itertools import groupby
from operator import itemgetter

def bodeguero(request):
    usuario = request.session.get('usuario')
    if not usuario or usuario.get('rol', '').lower() != 'bodeguero':
        messages.error(request, 'No tienes permisos para acceder a esta sección.')
        return redirect('/login/')
    try:
        response_usu = requests.get('http://127.0.0.1:8001/usuarios/usuarios/')
        clientes = response_usu.json()

        response_prod = requests.get('http://localhost:3000/productos')
        productos = response_prod.json() if response_prod.status_code == 200 else []

        response_pendientes = requests.get('http://localhost:3000/pedidos_bodega?estado=1')
        pedidos_pendientes = response_pendientes.json() if response_pendientes.status_code == 200 else []

        response_preparacion = requests.get('http://localhost:3000/pedidos_bodega?estado=2')
        pedidos_preparacion = response_preparacion.json() if response_preparacion.status_code == 200 else []

    except Exception as e:
        print("Error:", e)
        clientes, productos = [], []
        pedidos_pendientes, pedidos_preparacion = [], []

    return render(request, 'app/bodeguero.html', {
        'usuario': usuario,
        'clientes': clientes,
        'productos': productos,
        'pedidos_pendientes': pedidos_pendientes,
        'pedidos_preparacion': pedidos_preparacion
    })

def marcar_pedido_en_preparacion(request, id_pedido):
    if request.method == "POST":
        try:
            response = requests.put(
                f'http://localhost:3000/pedido_bodega/{id_pedido}',
                json={"nuevo_estado": 2}
            )
            response.raise_for_status()
        except requests.exceptions.RequestException as e:
            print("Error actualizando estado a 'en preparación':", e)
        return redirect('bodeguero')


def marcar_pedido_listo_entregar(request, id_pedido):
    if request.method == "POST":
        try:
            response = requests.put(
                f"http://localhost:3000/pedido_bodega/{id_pedido}",
                json={"nuevo_estado": 6}
            )
            response.raise_for_status()
        except requests.exceptions.RequestException as e:
            print("Error al marcar como listo para entregar:", e)
        return redirect('bodeguero')

def cliente(request):
    response = requests.get('http://127.0.0.1:8001/usuarios/usuarios')
    usuario = response.json()
    return render(request, 'app/cliente.html', {'usuario': usuario})


def contador(request):
    usuario = request.session.get('usuario')
    estado = request.GET.get('estado', 'N') 
    if not usuario or usuario.get('rol', '').lower() != 'contador':
        messages.error(request, 'No tienes permisos para acceder a esta sección.')
        return redirect('/login/')
    try:
        response = requests.get(f'http://localhost:3000/transferencia?estado={estado}')
        if response.status_code == 200:
            transferencias = response.json()
            print("Transferencias obtenidas:", transferencias)
        else:
            transferencias = []
    except Exception as e:
        transferencias = []
        print(f'Error al obtener transferencias: {e}')

    return render(request, 'app/contador.html', {
        'transferencias': transferencias,
        'estado_actual': estado
    })

def vendedor(request):
    usuario = request.session.get('usuario')
    if not usuario or usuario.get('rol', '').lower() != 'vendedor':
        messages.error(request, 'No tienes permisos para acceder a esta sección.')
        return redirect('/login/')
    try:
        response = requests.get('http://localhost:3000/pedidos_bodega?estado=6')  
        pedidos = response.json() if response.status_code == 200 else []
    except Exception as e:
            print("Error al obtener pedidos:", e)
            pedidos = []

    try:
        response_usu = requests.get('http://127.0.0.1:8001/usuarios/usuarios/')
        clientes = response_usu.json()

        response = requests.get('http://localhost:3000/productos')
        productos = response.json() if response.status_code == 200 else []
    except Exception as e:
        print("Error:", e)
        clientes, productos = [], []

    return render(request, 'app/vendedor.html', {
        'usuario': usuario,
        'clientes': clientes,
        'productos': productos,
        'pedidos': pedidos
    })

def actualizar_estado_pedido_vendedor(request, id_pedido, tipo_entrega):
    if request.method == "POST":
        nuevo_estado = 7 if tipo_entrega == 2 else 4

        try:
            response = requests.put(
                f"http://localhost:3000/pedido_bodega/{id_pedido}",
                json={"nuevo_estado": nuevo_estado}
            )
            response.raise_for_status()
        except Exception as e:
            print("Error al actualizar pedido:", e)

    return redirect("vendedor")

def catalogo(request):
    return render(request, 'app/catalogo.html')

def base(request):
    return render(request, 'app/base.html')

def login(request):
    if request.method == 'POST':
        email = request.POST.get('email')
        clave = request.POST.get('clave')

        try:
            response = requests.get(
                'http://127.0.0.1:8001/usuarios/login',
                params={'email': email, 'clave': clave}
            )

            print("Status code:", response.status_code)
            print("Response content:", response.text)

            if response.status_code == 200:
                usuario = response.json()
                print("Usuario recibido:", usuario)

                rol = usuario.get('rol', '').lower()
                requiere_cambio = usuario.get('requiere_cambio')

                request.session['usuario'] = usuario
                request.session['requiere_cambio'] = requiere_cambio
                request.session['usuario_rut'] = usuario.get('rut')
                request.session['usuario_rol'] = rol
                request.session['usuario_email'] = email
                request.session['usuario_nombre'] = usuario.get('nombre')

                if rol == 'administrador' and requiere_cambio == 1:
                    return redirect('cambiar_clave')

                redirecciones = {
                    'cliente': 'index',
                    'vendedor': 'vendedor',
                    'bodeguero': 'bodeguero',
                    'contador': 'contador',
                    'administrador': 'administrador'
                }
                if request.GET.get("logout") == "1":
                    messages.success(request, "Has cerrado sesión correctamente.")

                if rol in redirecciones:
                    messages.success(request, 'Inicio de sesión exitoso.')
                    return redirect(redirecciones[rol])
                else:
                    messages.error(request, 'Rol no reconocido.')
                    return redirect('login')

            else:
                messages.error(request, 'Email o clave incorrectos.')
                return redirect('login')

        except Exception as e:
            messages.error(request, f'Error del servidor: {e}')
            return redirect('login')

    return render(request, 'app/login.html')

def cambiar_clave(request):
    if request.method == 'POST':
        email = request.session.get('usuario_email')
        nueva_clave = request.POST.get('nueva_clave')
        repetir_clave = request.POST.get('repetir_clave')

        if not nueva_clave or not repetir_clave:
            messages.error(request, "Por favor, complete ambos campos de clave.")
            return redirect('cambiar_clave')

        if nueva_clave != repetir_clave:
            messages.error(request, "Las contraseñas no coinciden.")
            return redirect('cambiar_clave')

        
        try:
            response = requests.put(
                'http://127.0.0.1:8001/usuarios/cambiar-clave',
                json={"email": email, "nueva_clave": nueva_clave}
            )

            if response.status_code == 200:
                messages.success(request, "Clave actualizada correctamente. Por favor, inicie sesión nuevamente.")
           
                request.session.flush()
                return redirect('login')
            else:
                messages.error(request, f"Error al cambiar clave: {response.json().get('detail', 'Desconocido')}")
                return redirect('cambiar_clave')

        except Exception as e:
            messages.error(request, f"Error del servidor: {e}")
            return redirect('cambiar_clave')

    
    return render(request, 'app/cambiar_clave.html')

def mostrar_administradores(request):
    try:
        response = requests.get('http://127.0.0.1:8001/administrador')
        if response.status_code == 200:
            administradores = response.json()
        else:
            administradores = []
            messages.error(request, "Error al obtener administradores")
    except Exception as e:
        administradores = []
        messages.error(request, f"Error en el servidor: {e}")

    return render(request, 'app/administradores.html', {'administradores': administradores})

def crear_admin(request):
    if request.method == 'POST':
        rut = request.POST.get('rut')
        nombre = request.POST.get('nombre')
        email = request.POST.get('email')
        clave = request.POST.get('clave')
        

        try:
            response = requests.post(
                'http://127.0.0.1:8001/usuarios/registro-administrador',
                json={
                    'rut': rut,
                    'nombre': nombre,
                    'email': email,
                    'clave': clave,
                    'rol': 'administrador'
                }
            )

            if response.status_code == 200:
                messages.success(request, 'Administrador creado correctamente.')
                return redirect('login')
            else:
                messages.error(request, f'Error al crear el administrador: {response.text}')
        except Exception as e:
            messages.error(request, f'Error de conexión: {e}')

    return render(request, 'app/crear_admin.html')

def register(request):
    try:
        comuna_response = requests.get('http://127.0.0.1:8001/usuarios/comuna')
        comunas = comuna_response.json()
    except Exception as e:
        comunas = []
        messages.error(request, 'Error al cargar las comunas')

    if request.method == 'POST':
        rut = request.POST.get('rut')
        nombre = request.POST.get('nombre')
        apellido = request.POST.get('apellido')
        email = request.POST.get('email')
        telefono = request.POST.get('telefono')
        clave = request.POST.get('clave')
        id_comuna = request.POST.get('id_comuna')

        data = {
            'rut': rut,
            'nombre': nombre,
            'apellido': apellido,
            'email': email,
            'telefono': telefono,
            'clave': clave,
            'id_comuna': id_comuna
        }

        response = requests.post('http://127.0.0.1:8001/usuarios/registro-cliente', params=data)

        if response.status_code == 200:
            messages.success(request, 'Registrado con éxito')
            return redirect('/login/')
        else:
            messages.error(request, response.json().get('detail', 'Error al registrar cliente'))

    return render(request, 'app/register.html', {'comunas': comunas})

def administrador(request):
    usuario = request.session.get('usuario')
    if not usuario:
        return redirect('/login/')

    if usuario.get('rol', '').lower() != 'administrador':
        messages.error(request, 'No tienes permisos para acceder a esta sección.')
        return redirect('/login/')

    try:
        response = requests.get('http://127.0.0.1:8001/usuarios/usuarios/')
        if response.status_code == 200:
            todos_usuarios = response.json()

            trabajadores = [u for u in todos_usuarios if u['rol'].lower() in ['vendedor', 'contador', 'bodeguero']]
        else:
            trabajadores = []
            messages.error(request, 'No se pudieron obtener los usuarios desde la API.')
    except Exception as e:
        trabajadores = []
        messages.error(request, f'Error al conectar con la API: {e}')

    return render(request, 'app/admin/administrador.html', {
        'trabajadores': trabajadores,
        'usuario': usuario 
    })

def eliminar_trabajador(request, rut):
    if request.method == "POST":
        url = f"http://127.0.0.1:8001/usuarios/eliminar/{rut}" 
        try:
            response = requests.delete(url)
            if response.status_code == 200:
                messages.success(request, "Trabajador eliminado con éxito.")
            elif response.status_code == 404:
                messages.error(request, "El trabajador no fue encontrado en la API.")
            else:
                messages.error(request, "Hubo un error al eliminar el trabajador.")
        except requests.exceptions.RequestException as e:
            messages.error(request, f"Error de conexión con la API: {e}")
    
    return redirect('administrador')

def agregarTrab(request):
    if request.method == 'POST':
        telefono_str = request.POST.get('telefono')
        try:
            telefono = int(telefono_str)
        except ValueError:
            messages.error(request, 'El número de teléfono debe ser numérico.')
            return render(request, 'app/admin/agregarTrab.html')
        data = {
            'rut': request.POST.get('rut'),
            'nombre': request.POST.get('nombre'),
            'apellido': request.POST.get('apellido'),
            'email': request.POST.get('email'),
            'telefono': telefono,
            'clave': request.POST.get('clave'),
            'rol': request.POST.get('rol'),
            'id_comuna': request.POST.get('id_comuna') 
        }

        response = requests.post('http://127.0.0.1:8001/usuarios/registro-trabajador', json=data)
        if response.status_code == 200:
            messages.success(request, 'Trabajador registrado con éxito')
            return redirect('/administrador/')
        else:
            messages.error(request, response.json().get('detail', 'Error al registrar trabajador'))
            return render(request, 'app/admin/agregarTrab.html')
    
    # Si necesitas mostrar las comunas en el formulario:
    comunas = requests.get('http://127.0.0.1:8001/usuarios/comuna').json()  # Ajusta URL si es necesario
    return render(request, 'app/admin/agregarTrab.html', {'comunas': comunas})

def equiposMedicion(request):
    response = requests.get('http://localhost:3000/filtrar_productos/8')  
    try:
        productos = response.json()
    except ValueError:
        productos = []

    productos_convertidos = []
    if isinstance(productos, list):
        for p in productos:
            if isinstance(p, dict):
                productos_convertidos.append({k.lower(): v for k, v in p.items()})
    else:
        productos_convertidos = []
    return render(request, 'app/categorias/equiposMedicion.html', {'producto_filtrados': productos_convertidos})

def fijaciones_y_adhesivos(request):
    response = requests.get('http://localhost:3000/filtrar_productos/7')  
    try:
        productos = response.json()
    except ValueError:
        productos = []

    productos_convertidos = []
    if isinstance(productos, list):
        for p in productos:
            if isinstance(p, dict):
                productos_convertidos.append({k.lower(): v for k, v in p.items()})
    else:
        productos_convertidos = []

    return render(request, 'app/categorias/fijaciones_y_adhesivos.html',{'producto_filtrados': productos_convertidos})

def herramientasManu(request):
    response = requests.get('http://localhost:3000/filtrar_productos/1')  
    try:
        productos = response.json()
    except ValueError:
        productos = []

    productos_convertidos = []
    if isinstance(productos, list):
        for p in productos:
            if isinstance(p, dict):
                productos_convertidos.append({k.lower(): v for k, v in p.items()})
    else:
        productos_convertidos = []

    return render(request, 'app/categorias/herramientasManu.html', {'producto_filtrados': productos_convertidos})


def herramientasElec(request):
    response = requests.get('http://localhost:3000/filtrar_productos/2')  
    try:
        productos = response.json()
    except ValueError:
        productos = []

    productos_convertidos = []
    if isinstance(productos, list):
        for p in productos:
            if isinstance(p, dict):
                productos_convertidos.append({k.lower(): v for k, v in p.items()})
    else:
        productos_convertidos = []
    return render(request, 'app/categorias/herramientasElec.html',{'producto_filtrados': productos_convertidos})


def materialesBasic(request):
    response = requests.get('http://localhost:3000/filtrar_productos/3')
    try:
        productos = response.json()
    except ValueError:
        productos = []

    productos_convertidos = []
    if isinstance(productos, list):
        for p in productos:
            if isinstance(p, dict):
                productos_convertidos.append({k.lower(): v for k, v in p.items()})
    else:
        productos_convertidos = []

    return render(request, 'app/categorias/materialesBasic.html', {
        'producto_filtrados': productos_convertidos
    })


def tornillo_y_anclaje(request):
    response = requests.get('http://localhost:3000/filtrar_productos/6')  
    try:
        productos = response.json()
    except ValueError:
        productos = []

    productos_convertidos = []
    if isinstance(productos, list):
        for p in productos:
            if isinstance(p, dict):
                productos_convertidos.append({k.lower(): v for k, v in p.items()})
    else:
        productos_convertidos = []
    return render(request, 'app/categorias/tornillo_y_anclaje.html',{'producto_filtrados': productos_convertidos})


def equiposSeguridad(request):
    response = requests.get('http://localhost:3000/filtrar_productos/5')  
    try:
        productos = response.json()
    except ValueError:
        productos = []

    productos_convertidos = []
    if isinstance(productos, list):
        for p in productos:
            if isinstance(p, dict):
                productos_convertidos.append({k.lower(): v for k, v in p.items()})
    else:
        productos_convertidos = []
    return render(request, 'app/categorias/equiposSeguridad.html',{'producto_filtrados': productos_convertidos})


def detalle_producto(request, id_producto):
    response = requests.get(f'http://localhost:3000/productos/{id_producto}')
    producto = response.json()

    response_otros = requests.get('http://localhost:3000/productos')
    todos_los_productos = response_otros.json()
    otros_productos = [p for p in todos_los_productos if p["id_producto"] != id_producto][:8]
    return render(request, 'app/detalle_producto.html', {'producto': producto,'otros_productos': otros_productos})

def detalleTrab(request, rut_buscar):
    response = requests.get(f'http://127.0.0.1:8001/usuarios/usuario/{rut_buscar}')
    trabajador = response.json()
    return render(request, 'app/admin/detalleTrab.html', {'trabajador': trabajador})   

def modificar_trabajador(request, rut):
    # Obtener trabajador
    response = requests.get(f"http://127.0.0.1:8001/usuarios/usuario/{rut}")
    trabajador = response.json()
    trabajador = trabajador[0] if isinstance(trabajador, list) and trabajador else trabajador

    # Normalizar comuna
    trabajador['id_comuna'] = int(trabajador['id_comuna']) if trabajador.get('id_comuna') else None

    # Obtener comunas
    comunas = requests.get('http://127.0.0.1:8001/usuarios/comuna').json()

    if request.method == 'POST':
        data = {
            'nombre': request.POST.get('nombre'),
            'apellido': request.POST.get('apellido'),
            'email': request.POST.get('email'),
            'telefono': request.POST.get('telefono'),
            'rol': request.POST.get('rol'),
            'id_comuna': request.POST.get('id_comuna')
        }

        if data['id_comuna']:
            data['id_comuna'] = int(data['id_comuna'])

        data = {k: v for k, v in data.items() if v not in [None, '']}

        try:
            response = requests.patch(
                f"http://127.0.0.1:8001/usuarios/modificar/{rut}",
                json=data,
                headers={'Content-Type': 'application/json'}
            )

            if response.status_code == 200:
                messages.success(request, "Trabajador actualizado correctamente")
                return render(request, 'app/admin/detalleTrab.html', {
                    'trabajador': trabajador,
                    'comunas': comunas,
                    'redirigir': True
                })
            else:
                error_detail = response.json().get('detail', response.text)
                messages.error(request, f"Error al actualizar: {error_detail}")
        except Exception as e:
            messages.error(request, f"Error de conexión: {str(e)}")

    return render(request, 'app/admin/detalleTrab.html', {
        'trabajador': trabajador,
        'comunas': comunas
    })

def detalleCliente(request, rut_buscar):
    response = requests.get(f'http://127.0.0.1:8001/usuarios/usuario/{rut_buscar}')
    cliente = response.json()
    return render(request, 'app/detalleCliente.html', {'cliente': cliente})

def modificar_cliente(request, rut):
    try:
        response = requests.get(f"http://127.0.0.1:8001/usuarios/usuario/{rut}")
        
        if response.status_code != 200:
            messages.error(request, "El cliente no existe o hay un problema con el servidor")
            return redirect('index')
            
        cliente = response.json()
        
        if isinstance(cliente, list):
            if len(cliente) == 0:
                messages.error(request, "No se encontró el cliente")
                return redirect('index')
            cliente = cliente[0]
        
        if not cliente.get('rut'):
            messages.error(request, "El cliente no tiene RUT válido")
            return redirect('index')
            

        cliente['id_comuna'] = int(cliente['id_comuna']) if cliente.get('id_comuna') else None

        comunas_response = requests.get('http://127.0.0.1:8001/usuarios/comuna')
        if comunas_response.status_code != 200:
            messages.error(request, "Error al obtener las comunas")
            return redirect('index')
            
        comunas = comunas_response.json()

        if request.method == 'POST':
            data = {
                'nombre': request.POST.get('nombre'),
                'apellido': request.POST.get('apellido'),
                'email': request.POST.get('email'),
                'telefono': request.POST.get('telefono'),
                'id_comuna': request.POST.get('id_comuna')
            }

            if not all([data['nombre'], data['apellido'], data['email'], data['telefono']]):
                messages.error(request, "Todos los campos son obligatorios")
                return render(request, 'app/detalleCliente.html', {
                    'cliente': cliente,
                    'comunas': comunas
                })

            if data['id_comuna']:
                try:
                    data['id_comuna'] = int(data['id_comuna'])
                except ValueError:
                    messages.error(request, "Comuna inválida")
                    return render(request, 'app/detalleCliente.html', {
                        'cliente': cliente,
                        'comunas': comunas
                    })

            try:
                response = requests.patch(
                    f"http://127.0.0.1:8001/usuarios/modificar/{rut}",
                    json=data,
                    headers={'Content-Type': 'application/json'}
                )

                if response.status_code == 200:
                    messages.success(request, "Cliente actualizado correctamente")
                
                    if request.session.get('usuario_rut') == rut:
                        request.session['usuario_nombre'] = data['nombre']
                        request.session['usuario_apellido'] = data['apellido']
                    return redirect('detalle_cliente', rut_buscar=rut)
                else:
                    error_detail = response.json().get('detail', response.text)
                    messages.error(request, f"Error al actualizar: {error_detail}")
            except requests.exceptions.RequestException as e:
                messages.error(request, f"Error de conexión: {str(e)}")

        return render(request, 'app/detalleCliente.html', {
            'cliente': cliente,
            'comunas': comunas
        })

    except Exception as e:
        messages.error(request, f"Error inesperado: {str(e)}")
        return redirect('index')


def agregar_al_carrito(request, producto_id):
    usuario = request.session.get("usuario")
    if not usuario:
        return redirect("login")
    
    if request.method == "POST":
        cantidad = int(request.POST.get("cantidad", 1))
        producto_id = int(producto_id)

        carrito = request.session.get("carrito", [])
        if not isinstance(carrito, list):
            carrito = []

        for item in carrito:
            if int(item["id_producto"]) == producto_id:
                item["cantidad"] += cantidad
                break
        else:
            carrito.append({"id_producto": producto_id, "cantidad": cantidad})

        request.session["carrito"] = carrito
        return redirect("carrito")
        
    return redirect("detalle_producto", producto_id=producto_id)

@require_POST
def vaciar_carrito(request):
    request.session['carrito'] = []
    return JsonResponse({'status': 'ok'})

def carrito(request):
    usuario = request.session.get("usuario")
    if not usuario or usuario.get('rol', '').lower() != 'cliente':
        messages.error(request, 'No tienes permisos para acceder a esta sección.')
        return redirect('/login/')
    if not usuario:
        return redirect("login")
    
    carrito = request.session.get("carrito", [])
    productos = []
    total = 0
    cantidad_total_productos = 0

    for item in carrito:
        response = requests.get(f"http://localhost:3000/productos/{item['id_producto']}")
        producto = response.json()
        producto["cantidad"] = item["cantidad"]
        producto["subtotal"] = producto["precio_prod"] * item["cantidad"]
        total += producto["subtotal"]
        cantidad_total_productos += item["cantidad"]
        productos.append(producto)

    try:
        response_dolar = requests.get("https://mindicador.cl/api/dolar")
        response_dolar.raise_for_status()
        data_dolar = response_dolar.json()
        valor_dolar = data_dolar["serie"][0]["valor"]
        total_usd = round(total / valor_dolar, 2)
    except requests.RequestException as e:
        print("Error al obtener el dólar:", e)
        valor_dolar = None
        total_usd = None

    return render(request, "app/carrito.html", {
        "productos": productos,
        "total": total,
        "total_usd": total_usd,
        "valor_dolar": valor_dolar,
        "cantidad_total_productos": cantidad_total_productos
    })

@csrf_exempt
def eliminar_del_carrito(request, producto_id):
    if request.method == "POST":
        carrito = request.session.get("carrito", [])
        carrito = [item for item in carrito if item["id_producto"] != producto_id]
        request.session["carrito"] = carrito
    return redirect("carrito")


@csrf_exempt
def reducir_cantidad(request, producto_id):
    if request.method == "POST":
        carrito = request.session.get("carrito", [])
        for item in carrito:
            if item["id_producto"] == producto_id:
                item["cantidad"] -= 1
                if item["cantidad"] <= 0:
                    carrito = [i for i in carrito if i["id_producto"] != producto_id]
                break
        request.session["carrito"] = carrito
    return redirect("carrito")

@csrf_exempt
def aumentar_cantidad(request, producto_id):
    if request.method == "POST":
        carrito = request.session.get("carrito", [])
        for item in carrito:
            if item["id_producto"] == producto_id:
                item["cantidad"] += 1
                break
        request.session["carrito"] = carrito
    return redirect("carrito")

def logout(request):
    request.session.flush()
    return redirect("/login?logout=1")


@csrf_protect
def checkout(request):
    usuario = request.session.get("usuario")
    if not usuario:
        return redirect("login")

    carrito = request.session.get("carrito", [])
    productos = []
    total = 0
    cantidad_total_productos = 0
    mensaje_error = None
    descuento_aplicado = request.session.get('descuento_aplicado', False)
    cupon_codigo = request.session.get('cupon_codigo')
    descuento_valor = request.session.get('descuento_valor', 0)

    try:
        res_tipo = requests.get("http://localhost:3000/tipo_entrega")
        if res_tipo.status_code == 200:
            tipos_entrega = res_tipo.json()
            print("TIPOS ENTREGA:", tipos_entrega)
        else:
            tipos_entrega = []
            print("Error al obtener tipos de entrega")
    except Exception as e:
            tipos_entrega = []
            print("Error al conectar con la API de tipo_entrega:", str(e))

    try:
        res_suc = requests.get("http://localhost:3000/sucursales")
        if res_suc.status_code == 200:
            sucursales = res_suc.json()
            print("SUCURSALES:", sucursales)  
        else:
            sucursales = []
            print("Error al obtener sucursales. Código:", res_suc.status_code)
    except Exception as e:
        sucursales = []
        print("Error al conectar con la API de sucursales:", str(e))

    for item in carrito:
        try:
            response = requests.get(f"http://localhost:3000/productos/{item['id_producto']}")
            producto = response.json()
            producto["precio_prod"] = producto.get("precio_prod", 0)
        except Exception:
            producto = {"id_producto": item["id_producto"], "precio_prod": 0}

        producto["cantidad"] = item["cantidad"]
        producto["subtotal"] = producto["precio_prod"] * item["cantidad"]
        total += producto["subtotal"]
        cantidad_total_productos += item["cantidad"]
        productos.append(producto) 
    if request.method == "POST":
        codigo = request.POST.get("codigo")
        rut_usuario = usuario.get("rut")

        try:
            res = requests.get("http://localhost:3000/cupon/validar", params={
                "codigo": codigo,
                "rut": rut_usuario
            })

            if res.status_code == 200:
                data = res.json()
                request.session['cupon_codigo'] = data.get("codigo")
                request.session['descuento_aplicado'] = True
                request.session['descuento_valor'] = data.get("descuento", 0)
                descuento_aplicado = True
                descuento_valor = data.get("descuento", 0)
            else:
                mensaje_error = res.json().get("mensaje", "Cupón inválido o vencido.")
                request.session['descuento_aplicado'] = False
                request.session['descuento_valor'] = 0
        except Exception as e:
            mensaje_error = "Error al validar el cupón."
            print("Error al conectar con la API del cupón:", str(e))

    if descuento_aplicado:
        total = total - (total * descuento_valor / 100)

    request.session['total'] = total

    return render(request, "app/checkout.html", {
        "productos": productos,
        "total": total,
        "usuario": usuario,
        "sucursales": sucursales,
        "tipos_entrega": tipos_entrega,
        "mensaje_error": mensaje_error,
        "descuento_aplicado": descuento_aplicado,
        "cupon_codigo": cupon_codigo,
        "cantidad_total_productos": cantidad_total_productos,
})

def procesar_checkout(request):
    if request.method == "POST":
        request.session['checkout_data'] = {
            "rut": request.POST.get("rut"),
            "email": request.POST.get("email"),
            "telefono": request.POST.get("telefono"),
            "tipo_envio": request.POST.get("tipo_envio"),
            "direccion": request.POST.get("direccion"),
            "sucursal": request.POST.get("sucursal"),
            "cupon_codigo": request.POST.get("cupon_codigo"),
        }
        return redirect("paypalCompra") 
    return redirect("checkout")

from collections import defaultdict



def mis_pedidos(request):
    usuario = request.session.get("usuario")
    if not usuario:
        return redirect("login")

    rut = usuario.get("rut")
    pedidos_agrupados = defaultdict(lambda: {
        "productos": [],
        "total": 0,
        "estado": "",
        "fecha": "",
        "id_pedido": None,
        "tipo_pago": "Desconocido",
        "pago_confirmado": None
    })

    try:
        response = requests.get(f"http://localhost:3000/pedidos/{rut}")
        if response.status_code == 200:
            data = response.json()
            for item in data:
                item = {k.lower(): v for k, v in item.items()}
                id_pedido = item["id_pedido"]

                if id_pedido not in pedidos_agrupados:
                    pedidos_agrupados[id_pedido] = {
                        "id_pedido": id_pedido,
                        "fecha": item.get("fecha", "Sin fecha"),
                        "estado": item.get("estado_pedido") or "Desconocido",
                        "total": item.get("total", 0),
                        "tipo_pago": item.get("tipo_pago", "Desconocido"),
                        "pago_confirmado": item.get("pago_confirmado"),
                        "cupon_codigo": item.get("cupon_codigo"),
                        "porcentaje_descuento": item.get("porcentaje_descuento", 0),
                        "productos": []
                    }

                cantidad = item.get("cantidad") or 0
                precio_unitario = item.get("precio_unitario") or 0

                pedidos_agrupados[id_pedido]["productos"].append({
                    "nombre_prod": item.get("nombre_producto", "Producto"),
                    "cantidad": cantidad,
                    "precio_unitario": precio_unitario,
                    "subtotal": cantidad * precio_unitario,
                    "estado": item.get("estado_pedido") or "Desconocido"
                })

    except Exception as e:
        print("Error al obtener pedidos:", e)

    pedidos = list(pedidos_agrupados.values())
    return render(request, "app/mis_pedidos.html", {"pedidos": pedidos})



def paypalCompra(request):
    usuario = request.session.get("usuario")
    if not usuario:
        return redirect("login")

    carrito = request.session.get("carrito", [])
    productos = []
    total = 0
    cantidad_total_productos = 0

    for item in carrito:
        response = requests.get(f"http://localhost:3000/productos/{item['id_producto']}")
        producto = response.json()
        producto["cantidad"] = item["cantidad"]
        producto["subtotal"] = producto["precio_prod"] * item["cantidad"]
        total += producto["subtotal"]
        cantidad_total_productos += item["cantidad"]
        productos.append(producto)

    # Obtener cupón aplicado de la sesión
    cupon = request.session.get("cupon_aplicado")
    descuento = 0
    total_final = total

    if cupon:
        # Verificar estado actual del cupón desde la API
        response = requests.get(f"http://localhost:3000/cupones/{cupon['codigo']}")
        if response.status_code == 200:
            cupon_data = response.json()
            if not cupon_data.get("usado", False):
                descuento = total * (cupon_data["descuento"] / 100)
                total_final = total - descuento
            else:
                # Cupón ya fue usado, lo eliminamos de la sesión
                request.session.pop("cupon_aplicado", None)
                cupon = None
        else:
            # No se pudo verificar el cupón, lo eliminamos por seguridad
            request.session.pop("cupon_aplicado", None)
            cupon = None

    return render(request, "app/paypalCompra.html", {
        "productos": productos,
        "total": total,
        "descuento": descuento,
        "total_final": total_final,
        "cupon": cupon,
        "usuario": usuario,
        "cantidad_total_productos": cantidad_total_productos
})

def descargar_boleta(request, id_pedido):
    try:
        detalle_response = requests.get(f"http://localhost:3000/detalle_pedido/{id_pedido}")
        if detalle_response.status_code != 200:
            return HttpResponse("No se pudo obtener el detalle del pedido.", status=404)

        detalles = detalle_response.json()
        if not detalles:
            return HttpResponse("El pedido no tiene productos.", status=404)

        response = HttpResponse(content_type='application/pdf')
        response['Content-Disposition'] = f'attachment; filename="boleta_{id_pedido}.pdf"'

        p = canvas.Canvas(response, pagesize=letter)
        width, height = letter

        y = height - 30

        logo_path = os.path.join(settings.BASE_DIR, 'app','static', 'app', 'img', 'logo2.png')
        try:
            print("Ruta al logo:", logo_path)
            print("Existe:", os.path.exists(logo_path))
            p.drawImage(logo_path, 50, height - 100, width=120, preserveAspectRatio=True, mask='auto')
        except Exception as img_error:
            print("No se pudo cargar el logo:", img_error)
        
        p.setFont("Helvetica-Bold", 12)
        p.drawCentredString(width / 2, y, "R.U.T.: 76.192.083-9")
        y -= 15
        p.drawCentredString(width / 2, y, "BOLETA ELECTRÓNICA")
        y -= 15
        p.drawCentredString(width / 2, y, f"N° {id_pedido}")
        y -= 20
        p.drawCentredString(width / 2, y, "S.I.I. - Santiago")

        y -= 30
        p.setFont("Helvetica-Bold", 10)
        p.drawString(50, y, "Ferremas FRM")
        y -= 12
        p.setFont("Helvetica", 10)
        p.drawString(50, y, "FERRETERÍA FERREMAS")
        y -= 12
        p.drawString(50, y, "Santiago centro 191, Santiago")

        y -= 20
        fecha = datetime.now().strftime("%d de %B del %Y")
        p.drawString(50, y, f"Emisión   : {fecha}")

        y -= 30
        p.line(40, y, width - 40, y)
        y -= 15
        p.setFont("Helvetica-Bold", 10)
        p.drawString(50, y, "Item")
        p.drawString(250, y, "P. unitario")
        p.drawString(350, y, "Cant.")
        p.drawString(420, y, "Total item")

        y -= 5
        p.line(40, y, width - 40, y)
        y -= 15

        total_general = 0
        p.setFont("Helvetica", 10)

        for detalle in detalles:
            producto = detalle["nombre_prod"]
            precio = detalle["precio_unitario"]
            cantidad = detalle["cantidad"]
            subtotal = precio * cantidad
            total_general += subtotal

            p.drawString(50, y, producto[:30])
            p.drawRightString(300, y, f"${precio:,.0f}")
            p.drawRightString(390, y, f"{cantidad}")
            p.drawRightString(500, y, f"${subtotal:,.0f}")
            y -= 20

        p.line(40, y, width - 40, y)
        y -= 20

        # Obtener si hay descuento aplicado desde el pedido
        pedido_resp = requests.get(f"http://localhost:3000/pedido/{id_pedido}")
        if pedido_resp.status_code == 200:
            pedido_data = pedido_resp.json()
            cupon_aplicado = pedido_data.get("id_cupon") is not None
        else:
            cupon_aplicado = False  # Si falla la petición, asumimos sin cupón

        descuento_valor = total_general * 0.10 if cupon_aplicado else 0
        total_final = total_general - descuento_valor

        p.setFont("Helvetica-Bold", 10)
        p.drawRightString(500, y, f"Subtotal     : ${total_general:,.0f}")
        y -= 15
        if descuento_valor > 0:
            p.drawRightString(500, y, f"Descuento 10%: -${descuento_valor:,.0f}")
            y -= 15
        p.drawRightString(500, y, f"Total Final  : ${total_final:,.0f}")


        p.showPage()
        p.save()

        return response

    except Exception as e:
        print("Error al generar PDF:", e)
        return HttpResponse("Error interno al generar boleta", status=500)

@csrf_protect
def aplicar_cupon(request):
    if request.method == "POST":
        codigo = request.POST.get("codigo")
        usuario = request.session.get("usuario")

        if not usuario:
            return redirect("login")

        rut_usuario = usuario.get("rut")
        mensaje_error = None

        try:
            res = requests.get("http://localhost:3000/cupon/validar", params={
                "codigo": codigo,
                "rut": rut_usuario
            })
            if res.status_code == 200 and res.json().get("valido"):
                request.session['cupon_codigo'] = codigo
                request.session['descuento_aplicado'] = True
            else:
                mensaje_error = "Cupón inválido o vencido."
                request.session['descuento_aplicado'] = False
        except Exception as e:
            mensaje_error = "Error al validar el cupón."
            print("Error al conectar con la API del cupón:", str(e))

        if mensaje_error:
            messages.error(request, mensaje_error)
        else:
            messages.success(request, f"Cupón {codigo} aplicado correctamente.")

    return redirect("checkout")

# --- Función para generar código aleatorio ---
def generar_codigo_cupon(longitud=8):
    caracteres = string.ascii_uppercase + string.digits
    return ''.join(random.choices(caracteres, k=longitud))

# --- Función para guardar cupón en la base Oracle ---
def guardar_cupon_db(rut_usuario, email=None):
    url = "http://localhost:3000/cupon/crear"
    payload = {
        "rut_usuario": rut_usuario,
        "email": email
    }
    try:
        response = requests.post(url, json=payload)
        if response.status_code == 201:
            data = response.json()
            return data.get("codigo"), data.get("id_cupon")
        else:
            print(f"Error creando cupón: {response.status_code} {response.text}")
            return None, None
    except Exception as e:
        print(f"Excepción al crear cupón: {e}")
        return None, None

# --- Endpoint para crear cupón y guardarlo en sesión ---
@csrf_exempt
def crear_y_guardar_cupon(request):
    if request.method == "POST":
        try:
            data = json.loads(request.body)
            rut = data.get("rut")
            if not rut:
                return JsonResponse({"error": "Falta rut"}, status=400)

            codigo = generar_codigo_cupon()

            # Opcional: definir fecha expiración, por ejemplo 30 días después
            import datetime
            fecha_expiracion = datetime.date.today() + datetime.timedelta(days=30)

            guardar_cupon_db(codigo, rut, monto_minimo=None, descuento=10, fecha_expiracion=fecha_expiracion)

            # Guardar en sesión para usar luego en finalizar_pago_paypal
            request.session["cupon"] = {"codigo": codigo, "rut": rut}

            return JsonResponse({"success": True, "codigo": codigo})

        except Exception as e:
            return JsonResponse({"error": str(e)}, status=500)

    return JsonResponse({"error": "Método no permitido"}, status=405)

@csrf_exempt
def finalizar_pago_paypal(request):
    if request.method == "POST":
        try:
            data = json.loads(request.body)
            print("DATOS RECIBIDOS:", data)

            order_id = data.get("orderID")

            if request.session.get("last_order_id") == order_id:
                print("Pedido ya fue procesado anteriormente con este orderID")
                return JsonResponse({"error": "Pedido ya procesado"}, status=400)
            
            request.session["last_order_id"] = order_id

            usuario = request.session.get("usuario")
            print("USUARIO:", usuario)
            rut = usuario.get("rut")
            email = usuario.get("email")

            checkout_data = request.session.get("checkout_data", {})
            rut = checkout_data.get("rut")
            tipo_envio = checkout_data.get("tipo_envio")
            sucursal = checkout_data.get("sucursal")
            print("RUT:", rut, "TIPO ENVÍO:", tipo_envio, "SUCURSAL:", sucursal)

            carrito = request.session.get("carrito", [])
            print("CARRITO:", carrito)

            if not carrito:
                return JsonResponse({"error": "Carrito vacío"}, status=400)

            total = 0
            productos = []
            cupon_generado = False
            cantidad_total_productos = 0
            cumple_condicion_producto_igual = False

            for item in carrito:
                response = requests.get(f"http://localhost:3000/productos/{item['id_producto']}")
                producto = response.json()
                producto["cantidad"] = item["cantidad"]
                producto["precio_unitario"] = producto["precio_prod"]
                producto["subtotal"] = producto["precio_unitario"] * item["cantidad"]
                total += producto["subtotal"]
                producto["id_estado_pedido"] = 1
                productos.append(producto)

                cantidad_total_productos += 1
                if item["cantidad"] > 4:
                    cumple_condicion_producto_igual = True

            # Cupón aplicado desde sesión
            cupon_sesion = request.session.get("cupon_aplicado")
            codigo_cupon = cupon_sesion.get("codigo") if cupon_sesion else None

            # Si no hay cupón aplicado, y se cumplen condiciones de regalo
            if not codigo_cupon and (cantidad_total_productos > 4 or cumple_condicion_producto_igual):
                codigo_cupon = ''.join(random.choices(string.ascii_uppercase + string.digits, k=8))
                fecha_expiracion = date.today() + timedelta(days=30)

                crear_cupon_resp = requests.post(
                    "http://localhost:3000/cupon/crear",
                    json={
                        "rut_usuario": rut,
                        "email": email,
                        "codigo": codigo_cupon,
                        "descuento": 10,
                        "fecha_expiracion": fecha_expiracion.isoformat()
                    }
                )

                if crear_cupon_resp.status_code == 201:
                    request.session["cupon"] = {"codigo": codigo_cupon, "rut": rut}
                    cupon_generado = True
                else:
                    print(f"Error al crear cupón: {crear_cupon_resp.text}")
                    codigo_cupon = None

            descuento_valor = 0
            if codigo_cupon:
                descuento_valor = total * 0.10
                total -= descuento_valor

            pedido_payload = {
                'total': total,
                'rut': rut,
                'id_tipo_entrega': tipo_envio,
                'id_sucursal': int(sucursal) if tipo_envio == 'retiro' else None,
                'id_cupon': cupon_sesion.get("id") if cupon_sesion else None,
                'productos': [
                    {
                        "id_producto": p["id_producto"],
                        "cantidad": p["cantidad"],
                        "precio_unitario": p["precio_unitario"],
                        "subtotal": p["subtotal"],
                        "id_estado_pedido": p["id_estado_pedido"]
                    }
                    for p in productos
                ],
                'descuento_aplicado': descuento_valor,
                'id_tipo_pago': 1,
                'pago_confirmado': 'S'
            }

            print("ENVIANDO PEDIDO A API:", pedido_payload)
            pedido_resp = requests.post("http://localhost:3000/pedido", json=pedido_payload)
            print("RESPUESTA DE API PEDIDO:", pedido_resp.status_code, pedido_resp.text)

            if pedido_resp.status_code != 201:
                return JsonResponse({"error": "Error al crear pedido", "detalle": pedido_resp.text}, status=500)

            pedido_data = pedido_resp.json()
            id_pedido = pedido_data.get("id_pedido") or pedido_data.get("id")
            print("ID_PEDIDO:", id_pedido)

            # Marcar cupón como usado si se aplicó uno manual
            if cupon_sesion:
                try:
                    response = requests.patch(
                        f"http://localhost:3000/cupones/{cupon_sesion['codigo']}/usar"
                    )
                    if response.status_code == 200:
                        print("Cupón marcado como usado correctamente.")
                        request.session.pop("cupon_aplicado", None)
                    else:
                        print(f"Error al marcar cupón como usado: {response.status_code} {response.text}")
                except Exception as e:
                    print(f"Excepción al marcar cupón como usado: {e}")

            request.session["carrito"] = []

            # Enviar correo si se generó un nuevo cupón
            if email and cupon_generado:
                try:
                    asunto = "¡Gracias por tu compra! Aquí tienes tu cupón de descuento"
                    mensaje = f"""Hola,

Gracias por tu compra.

Como agradecimiento, te regalamos un cupón de descuento del 10% para tu próxima compra.

🎁 CÓDIGO DE CUPÓN: {codigo_cupon}

📅 Validez: 30 días desde hoy.

¡Te esperamos pronto!
"""
                    send_mail(
                        asunto,
                        mensaje,
                        settings.DEFAULT_FROM_EMAIL,
                        [email],
                        fail_silently=False
                    )
                except Exception as e:
                    print(f"Error enviando correo: {e}")

            return JsonResponse({"success": True})

        except Exception as e:
            print("Error al finalizar pedido PayPal:", str(e))
            return JsonResponse({"error": str(e)}, status=500)

    return JsonResponse({"error": "Método no permitido"}, status=405)
@require_POST
def eliminar_cupon(request):
    if 'cupon_codigo' in request.session:
        del request.session['cupon_codigo']
    return redirect('checkout') 

def validar_cupon(request):
    codigo = request.GET.get("codigo")
    rut = request.GET.get("rut")

    if not codigo or not rut:
        return JsonResponse({"error": "Código y RUT son obligatorios"}, status=400)

    try:
        response = requests.get(f"http://localhost:3000/cupon/validar?codigo={codigo}&rut={rut}")

        if response.status_code == 200:
            data = response.json()
            # Guardamos el descuento en sesión para usarlo en el pago
            request.session["cupon_aplicado"] = {
                "id": data["id"],
                "codigo": data["codigo"],
                "descuento": data["descuento"]
            }
            return JsonResponse({"mensaje": "Cupón válido", "descuento": data["descuento"]})
        else:
            return JsonResponse({"error": "Cupón inválido o expirado"}, status=400)

    except Exception as e:
        return JsonResponse({"error": str(e)}, status=500)
    

@csrf_exempt
def guardar_cupon_sesion(request):
    if request.method == "POST":
        try:
            data = json.loads(request.body)
            codigo = data.get("codigo")
            rut = data.get("rut")
            if not codigo or not rut:
                return JsonResponse({"error": "Datos incompletos"}, status=400)

            # Guardar cupón en sesión para usar luego en finalizar_pago_paypal
            request.session["cupon"] = {"codigo": codigo, "rut": rut}
            return JsonResponse({"success": True})

        except Exception as e:
            return JsonResponse({"error": str(e)}, status=500)

    return JsonResponse({"error": "Método no permitido"}, status=405)



def generar_cupon(longitud=8):
    caracteres = string.ascii_uppercase + string.digits
    return ''.join(random.choices(caracteres, k=longitud))

def pago_transferencia(request):
    if request.method == 'POST' and request.FILES.get('comprobante'):
        comprobante = request.FILES['comprobante']
        fs = FileSystemStorage(location='media/comprobantes')
        filename = fs.save(comprobante.name, comprobante)
        comprobante_url = fs.url(filename)

        usuario = request.session.get('usuario')
        rut = usuario.get('rut')
        nombre_cliente = usuario.get('nombre', 'Invitado')
        email_cliente = usuario.get('email', None)
        tipo_envio = request.POST.get('tipo_envio')
        sucursal = request.POST.get('sucursal')
        carrito = request.session.get('carrito', [])

        total_final_str = request.POST.get('total_final')
        try:
            total_final = float(total_final_str)
        except (TypeError, ValueError):
            total_final = 0

        productos = []
        total = 0
        cantidad_total_productos = 0

        for item in carrito:
            response = requests.get(f"http://localhost:3000/productos/{item['id_producto']}")
            producto = response.json()
            producto["cantidad"] = item["cantidad"]
            producto["precio_unitario"] = producto["precio_prod"]
            producto["subtotal"] = producto["precio_unitario"] * item["cantidad"]
            producto["id_estado_pedido"] = 1  # Estado pendiente
            total += producto["subtotal"]
            cantidad_total_productos += item["cantidad"]
            productos.append(producto)

        cupon_sesion = request.session.get("cupon_aplicado")
        codigo_cupon = cupon_sesion.get("codigo") if cupon_sesion else None

        productos_distintos = set([item['id_producto'] for item in carrito])
        cantidad_productos_distintos = len(productos_distintos)

        generar_nuevo_cupon = (
            not codigo_cupon and 
            (cantidad_total_productos > 4 or cantidad_productos_distintos > 4)
        )

        if generar_nuevo_cupon:
            from datetime import date, timedelta
            codigo_cupon = generar_cupon()
            fecha_expiracion = date.today() + timedelta(days=30)

            crear_cupon_resp = requests.post(
                "http://localhost:3000/cupon/crear",
                json={
                    "rut_usuario": rut,
                    "email": email_cliente,
                    "codigo": codigo_cupon,
                    "descuento": 10,
                    "fecha_expiracion": fecha_expiracion.isoformat()
                }
            )
            if crear_cupon_resp.status_code == 201:
                request.session["cupon"] = {"codigo": codigo_cupon, "rut": rut}
            else:
                print(f"Error al crear cupón: {crear_cupon_resp.text}")
                codigo_cupon = None



        descuento_valor = 0
        if codigo_cupon:
            descuento_valor = total * 0.10
            total -= descuento_valor

        TIPOS_ENVIO = {'retiro': 1, 'domicilio': 2}
        id_tipo_entrega = TIPOS_ENVIO.get(tipo_envio, 2)

        # Crear payload para enviar a la API de pedido
        pedido_payload = {
            'total': total,
            'rut': rut,
            'id_tipo_entrega': id_tipo_entrega,
            'id_sucursal': int(sucursal) if tipo_envio == 'retiro' else None,
            'id_cupon': cupon_sesion.get("id") if cupon_sesion else None,
            'productos': [
                {
                    "id_producto": p["id_producto"],
                    "cantidad": p["cantidad"],
                    "precio_unitario": p["precio_unitario"],
                    "subtotal": p["subtotal"],
                    "id_estado_pedido": p["id_estado_pedido"]
                }
                for p in productos
            ],
            'id_tipo_pago': 3,  # Transferencia
            'pago_confirmado': 'N'
        }
        

        try:
            # Enviar el pedido a la API
            response_pedido = requests.post('http://localhost:3000/pedido', json=pedido_payload)
            if response_pedido.status_code == 201:
                data_pedido = response_pedido.json()
                id_pago = data_pedido.get('id_pago')  

                # ENVIAR CUPÓN SI COMPRA MÁS DE 4 PRODUCTOS
                if cantidad_total_productos > 4 and email_cliente and codigo_cupon:
                    cupon = generar_cupon()
                    asunto = "¡Gracias por tu compra! Aquí tienes un cupón de descuento"
                    mensaje = f"Hola {nombre_cliente},\n\nGracias por comprar en Ferremas. " \
                              f"Como agradecimiento, te enviamos este cupón de descuento: {codigo_cupon}\n\n" \
                              "¡Nos encantaría verte pronto de nuevo!"
                    send_mail(
                        asunto,
                        mensaje,
                        'no-reply@ferremas.cl',  # Remitente
                        [email_cliente],
                        fail_silently=False,
                    )
                  

                with fs.open(filename, 'rb') as f:
                    files = {'archivo_comprobante': (filename, f, comprobante.content_type)}
                    data = {
                        'rut': rut,
                        'total': total,
                        'id_pago': id_pago  
                    }
                    response_transferencia = requests.post(
                        'http://localhost:3000/transferencia/', data=data, files=files
                    )

                    if response_transferencia.status_code in [200, 201]:
                        request.session['carrito'] = {}
                        messages.success(request, 'Transferencia enviada al contador.')
                        return redirect('index')
                    else:
                        messages.error(request, f'Error al enviar comprobante: {response_transferencia.text}')
                        return redirect('index')
            else:
                messages.error(request, f'Error al registrar pedido: {response_pedido.text}')
                return redirect('index')
        except Exception as e:
            messages.error(request, f'Error de conexión: {e}')
            return redirect('index')

    messages.error(request, 'Debe subir un comprobante.')
    return redirect('index')


def confirmar_transferencia(request, id):
    try:
        requests.put(f'http://localhost:3000/transferencia/{id}/confirmar')
    except Exception as e:
        print(f'Error al confirmar: {e}')
    return redirect('contador')

def rechazar_transferencia(request, id):
    try:
        requests.put(f'http://localhost:3000/transferencia/{id}/rechazar')
    except Exception as e:
        print(f'Error al rechazar: {e}')
    return redirect('contador')

from django.core.mail import send_mail
    
def recuperar_contrasena(request):
    if request.method == "POST":
        email = request.POST.get("email")
        resp = requests.get("http://localhost:8001/usuarios/buscar-usuario-por-email/", params={"email": email})

        if resp.status_code == 200:
            rut = resp.json()["rut"]
            reset_link = request.build_absolute_uri(
            reverse('restablecer_contrasena', kwargs={'rut': rut})
                )

            send_mail(
                'Restablecer contraseña',
                f'Haz clic en el siguiente enlace para restablecer tu contraseña: {reset_link}',
                settings.DEFAULT_FROM_EMAIL,
                [email],
                fail_silently=False,
            )
            return render(request, "app/correo_enviado.html")

        return render(request, "app/recuperar_contrasena.html", {"error": "Correo no registrado"})
    
    return render(request, "app/recuperar_contrasena.html")

def restablecer_contrasena(request, rut):
    if request.method == "POST":
        nueva_clave = request.POST.get("nueva")
        confirmar_clave = request.POST.get("confirmar")

        if nueva_clave != confirmar_clave:
            return render(request, "app/restablecer.html", {
                "error": "Las contraseñas no coinciden",
                "rut": rut
            })

        try:
            response = requests.patch(
                f"http://localhost:8001/usuarios/modificar-clave/{rut}",
                json={"clave": nueva_clave},
                headers={'Content-Type': 'application/json'}
            )

            if response.status_code == 200:
                return render(request, "app/contrasena_actualizada.html")
            else:
                error_detail = response.json().get('detail', 'Error desconocido')
                return render(request, "app/restablecer.html", {
                    "error": f"Error del servidor: {error_detail}",
                    "rut": rut
                })
                
        except Exception as e:
            return render(request, "app/restablecer.html", {
                "error": f"Error de conexión: {str(e)}",
                "rut": rut
            })

    return render(request, "app/restablecer.html", {"rut": rut})

def productos_por_marca(request, id_marca):
    url = f'http://localhost:3000/productos/marca/{id_marca}'
    response = requests.get(url)
    productos = response.json()
    return render(request, 'app/productos_marca.html', {
        'productos': productos
    })

def informes(request):
    pagos = requests.get("http://localhost:3000/informes/pagos_realizados").json()
    pedidos = requests.get("http://localhost:3000/informes/pedidos_recibidos").json()
    productos = requests.get("http://localhost:3000/productos").json()

    return render(request, "app/informes.html", {
        "pagos": pagos,
        "pedidos": pedidos,
        "productos": productos
    })


def descargar_excel_pagos(request):
    pagos = requests.get("http://localhost:3000/informes/pagos_realizados").json()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Pagos Realizados"

    encabezados = ["ID Pedido", "Fecha", "RUT", "Total", "Tipo de Pago", "Producto", "Cantidad", "Precio Unitario", "Subtotal"]
    ws.append(encabezados)

    for p in pagos:
        ws.append([
            p["id_pedido"], p["fecha"], p["rut"], p["total"],
            p["tipo_pago"], p["nombre_producto"], p["cantidad"],
            p["precio_unitario"], p["subtotal"]
        ])

    response = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    response["Content-Disposition"] = "attachment; filename=pagos_realizados.xlsx"
    wb.save(response)
    return response


def descargar_excel_pedidos(request):
    pedidos = requests.get("http://localhost:3000/informes/pedidos_recibidos").json()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Pedidos Recibidos"

    encabezados = ["ID Pedido", "Fecha", "RUT", "Total", "Producto", "Cantidad", "Precio Unitario", "Subtotal", "Estado", "Pago Confirmado", "Tipo de Pago"]
    ws.append(encabezados)

    for p in pedidos:
        ws.append([
            p["id_pedido"], p["fecha"], p["rut"], p["total"],
            p["nombre_producto"], p["cantidad"], p["precio_unitario"],
            p["subtotal"], p["estado_pedido"], p["pago_confirmado"],
            p["tipo_pago"]
        ])

    response = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    response["Content-Disposition"] = "attachment; filename=pedidos_recibidos.xlsx"
    wb.save(response)
    return response

def descargar_excel_productos(request):
    response = requests.get("http://localhost:3000/productos")
    productos = response.json()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Productos"

    encabezados = [
        "ID Producto", "Código", "Modelo", "Marca", "Nombre", 
        "Descripción", "Precio", "Stock", "Subcategoría", "ID Subcategoría"
    ]
    ws.append(encabezados)

    for p in productos:
        ws.append([
            p["id_producto"],
            p["cod_producto"],
            p["modelo_prod"],
            p["nombre"],             # Marca
            p["nombre_prod"],
            p["descripcion"],
            p["precio_prod"],        # Ya viene como string formateado
            p["stock"],
            p["nombre_subcat"],
            p["id_subcategoria"]
        ])

    response = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    response["Content-Disposition"] = "attachment; filename=productos.xlsx"
    wb.save(response)
    return response



def eliminar_pedido(request, id_pedido):
    if request.method == 'POST':
        try:
            response = requests.delete(f'http://localhost:3000/pedido/{id_pedido}') 
            if response.status_code == 200:
                print("Pedido eliminado")
            else:
                print("Error al eliminar pedido:", response.status_code, response.text)
        except Exception as e:
            print("Excepción al eliminar pedido:", e)
    return redirect('mis_pedidos')

